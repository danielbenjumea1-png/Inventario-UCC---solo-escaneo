
import io
from flask import Flask, request, send_file, jsonify
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

app = Flask(__name__)

@app.route("/api/cruzar_inventario", methods=["POST"])
def cruzar_inventario():
    if "inventario" not in request.files or "escaneo" not in request.files:
        return jsonify({"error": "Faltan archivos"}), 400

    inventario_file = request.files["inventario"]
    escaneo_file = request.files["escaneo"]

    try:
        # Leer Excel con pandas
        df_inventario = pd.read_excel(inventario_file)
        df_escaneo = pd.read_excel(escaneo_file)

        # Normalizar columna 'codigo'
        df_inventario['codigo'] = df_inventario['codigo'].astype(str).str.strip()
        df_escaneo['codigo'] = df_escaneo['codigo'].astype(str).str.strip()

        # Crear set de códigos escaneados
        codigos_escaneo = set(df_escaneo['codigo'])

        # Guardar inventario temporalmente en memoria
        output_buffer = io.BytesIO()
        df_inventario.to_excel(output_buffer, index=False)
        output_buffer.seek(0)

        # Abrir con openpyxl para marcar en verde
        wb = load_workbook(output_buffer)
        ws = wb.active

        # Color verde
        verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

        # Buscar columna 'codigo'
        col_codigo = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "codigo":
                col_codigo = col
                break
        if col_codigo is None:
            return jsonify({"error": "No se encontró la columna 'codigo'"}), 400

        # Marcar coincidencias exactas
        coincidencias = 0
        for row in range(2, ws.max_row + 1):
            codigo = str(ws.cell(row=row, column=col_codigo).value).strip()
            if codigo in codigos_escaneo:
                ws.cell(row=row, column=col_codigo).fill = verde
                coincidencias += 1

        # Guardar Excel final en memoria
        final_buffer = io.BytesIO()
        wb.save(final_buffer)
        final_buffer.seek(0)

        # Devolver Excel como attachment + header con coincidencias
        return send_file(
            final_buffer,
            as_attachment=True,
            download_name="inventario_cruzado.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"X-Coincidencias": str(coincidencias)}
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == "__main__":
    app.run(debug=True)
